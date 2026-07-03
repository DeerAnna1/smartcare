"""Upload and media handling endpoints."""

import json
import secrets
from io import BytesIO
from datetime import datetime, timezone
from pathlib import Path

from fastapi import APIRouter, Depends, File, HTTPException, Request, UploadFile, status
from fastapi.responses import FileResponse
from sqlalchemy.ext.asyncio import AsyncSession
import httpx

from app.api.deps.auth import get_current_user, get_current_user_required
from app.core.config import get_settings
from app.core.database import get_db
from app.core.rate_limit import limiter
from app.models.models import User, LabReport
from app.services.lab_interpreter import parse_lab_text, summarize_lab_items

router = APIRouter(prefix="/upload", tags=["upload"])
settings = get_settings()

# parents[3] = /app  (mounted volume), so uploads persist across container restarts
UPLOAD_DIR = Path(__file__).resolve().parents[3] / "uploads"
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

ALLOWED_DOCUMENT_TYPES = {
    "application/pdf",
    "application/msword",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
    "text/plain",
    "image/jpeg",
    "image/png",
    "image/webp",
}
ALLOWED_IMAGE_TYPES = {"image/jpeg", "image/png", "image/webp"}
ALLOWED_DOCUMENT_EXTENSIONS = {".pdf", ".doc", ".docx", ".txt", ".jpg", ".jpeg", ".png", ".webp", ".xlsx", ".xls"}
ALLOWED_IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp"}

MAX_FILE_SIZE = 50 * 1024 * 1024  # 50MB for documents
MAX_IMAGE_SIZE = 10 * 1024 * 1024  # 10MB for images
MAX_EXTRACT_CHARS = 8000


def _safe_name(name: str) -> str:
    """Normalize user-supplied filename to prevent path traversal."""
    return Path(name).name.replace(" ", "_")


def _safe_json_loads(raw: str | None) -> dict:
    try:
        data = json.loads(raw or "{}")
        return data if isinstance(data, dict) else {}
    except json.JSONDecodeError:
        return {}


def _extension(name: str) -> str:
    return Path(name).suffix.lower()


def _extract_text_with_unstructured(content: bytes, filename: str) -> tuple[str, str]:
    """使用 unstructured 库解析文档（支持复杂 PDF 表格、OCR 等）。"""
    import tempfile
    import os
    try:
        from unstructured.partition.auto import partition
        # unstructured 需要文件路径，写入临时文件
        with tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(filename)[1]) as tmp:
            tmp.write(content)
            tmp_path = tmp.name
        try:
            elements = partition(filename=tmp_path)
            text_parts = []
            for elem in elements:
                text = str(elem).strip()
                if text:
                    text_parts.append(text)
            text = "\n".join(text_parts).strip()
            if not text:
                return "", "empty"
            return text[:MAX_EXTRACT_CHARS], "success"
        finally:
            os.unlink(tmp_path)
    except ImportError:
        return "", "failed"
    except Exception:
        return "", "failed"


def _extract_text_from_document(content: bytes, ext: str) -> tuple[str, str]:
    """Extract plain text from supported file types.

    Returns:
      - extracted text (possibly empty)
      - extraction status: success / unsupported / failed / empty
    """
    try:
        if ext in (".txt", ".md"):
            text = content.decode("utf-8", errors="ignore").strip()
            if not text:
                return "", "empty"
            return text[:MAX_EXTRACT_CHARS], "success"

        if ext == ".pdf":
            # 优先使用 unstructured（支持表格、OCR、复杂排版）
            text, status = _extract_text_with_unstructured(content, "document.pdf")
            if status == "success":
                return text, status
            # 回退到 pypdf
            from pypdf import PdfReader
            reader = PdfReader(BytesIO(content))
            parts: list[str] = []
            for page in reader.pages:
                try:
                    parts.append(page.extract_text() or "")
                except Exception:
                    parts.append("")
            text = "\n".join(parts).strip()
            if not text:
                return "", "empty"
            return text[:MAX_EXTRACT_CHARS], "success"

        if ext == ".docx":
            # 优先使用 unstructured
            text, status = _extract_text_with_unstructured(content, "document.docx")
            if status == "success":
                return text, status
            # 回退到 python-docx
            from docx import Document
            doc = Document(BytesIO(content))
            text = "\n".join(p.text for p in doc.paragraphs if p.text).strip()
            if not text:
                return "", "empty"
            return text[:MAX_EXTRACT_CHARS], "success"

        if ext in (".xlsx", ".xls"):
            import openpyxl
            wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
            parts: list[str] = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                parts.append(f"[{sheet}]")
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c) if c is not None else "" for c in row]
                    if any(cells):
                        parts.append("\t".join(cells))
            wb.close()
            text = "\n".join(parts).strip()
            if not text:
                return "", "empty"
            return text[:MAX_EXTRACT_CHARS], "success"

        # Legacy .doc upload is allowed, but text extraction is not supported here.
        return "", "unsupported"
    except Exception:
        return "", "failed"


async def _extract_text_via_ocr_space(content: bytes, filename: str) -> tuple[str, str]:
    """Use OCR.space API for OCR extraction."""
    if not settings.OCR_API_KEY:
        return "", "failed"
    try:
        form = {
            "language": "chs",
            "isOverlayRequired": "false",
            "OCREngine": "2",
            "isTable": "true",
            "scale": "true",
        }
        files = {"file": (filename, content)}
        headers = {"apikey": settings.OCR_API_KEY}
        async with httpx.AsyncClient(timeout=40) as client:
            resp = await client.post(
                "https://api.ocr.space/parse/image",
                data=form,
                files=files,
                headers=headers,
            )
        resp.raise_for_status()
        payload = resp.json()
        if payload.get("IsErroredOnProcessing"):
            return "", "failed"
        parsed = payload.get("ParsedResults") or []
        text = "\n".join((x.get("ParsedText") or "").strip() for x in parsed).strip()
        if not text:
            return "", "empty"
        return text[:MAX_EXTRACT_CHARS], "success"
    except Exception:
        return "", "failed"


async def _extract_text_via_llm_vision(content: bytes, ext: str) -> tuple[str, str]:
    """Use LLM vision model (mimo-v2-omni) to extract text from images."""
    try:
        import base64 as _b64
        from openai import AsyncOpenAI

        client = AsyncOpenAI(
            api_key=settings.OPENAI_API_KEY,
            base_url=settings.OPENAI_BASE_URL,
        )
        mime_map = {".jpg": "image/jpeg", ".jpeg": "image/jpeg", ".png": "image/png", ".webp": "image/webp"}
        mime = mime_map.get(ext, "image/png")
        img_url = f"data:{mime};base64,{_b64.b64encode(content).decode()}"

        # 使用多模态模型 mimo-v2-omni 进行图片分析
        response = await client.chat.completions.create(
            model=settings.MIMO_OMNI_MODEL,
            messages=[{
                "role": "user",
                "content": [
                    {"type": "text", "text": "请识别这张图片中的所有文字内容，包括表格、数字、标题等，原样输出。如果是医学检验报告，请保留所有数值和单位。只输出识别到的文字，不要添加任何解释。"},
                    {"type": "image_url", "image_url": {"url": img_url}}
                ]
            }],
            max_tokens=2000,
        )
        text = (response.choices[0].message.content or "").strip()
        if not text:
            return "", "empty"
        return text[:MAX_EXTRACT_CHARS], "success"
    except Exception:
        return "", "failed"


@router.post("/document")
@limiter.limit("10/minute")
async def upload_document(
    request: Request,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user_required),
):
    """Upload a document (PDF, Word, TXT)."""
    original_name = _safe_name(file.filename or "document")
    ext = _extension(original_name)
    if ext not in ALLOWED_DOCUMENT_EXTENSIONS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="文件类型不支持，仅支持 PDF/Word/TXT/图片",
        )

    if file.content_type and file.content_type not in ALLOWED_DOCUMENT_TYPES:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="文件内容类型不支持，仅支持 PDF/Word/TXT/图片",
        )

    content = await file.read()
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"文件超过 {MAX_FILE_SIZE // 1024 // 1024}MB 限制",
        )

    try:
        user_dir = UPLOAD_DIR / str(current_user.id) / "documents"
        user_dir.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        filename = f"{timestamp}_{secrets.token_hex(4)}_{original_name}"
        filepath = user_dir / filename

        with open(filepath, "wb") as f:
            f.write(content)

        # OCR 优先级：OCR.space → LLM 视觉模型 → 本地解析
        extracted_text = ""
        extraction_status = "failed"
        IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".webp"}

        if settings.OCR_PROVIDER.lower() == "ocr_space" and ext in (IMAGE_EXTS | {".pdf"}):
            extracted_text, extraction_status = await _extract_text_via_ocr_space(content, original_name)

        # 图片文件：OCR.space 失败时用 LLM 视觉模型识别
        if extraction_status != "success" and ext in IMAGE_EXTS:
            extracted_text, extraction_status = await _extract_text_via_llm_vision(content, ext)

        # 非图片或以上都失败时，用本地解析
        if extraction_status != "success":
            extracted_text, extraction_status = _extract_text_from_document(content, ext)
        structured_items = parse_lab_text(extracted_text) if extracted_text else []
        summary = summarize_lab_items(structured_items) if structured_items else ""
        report = LabReport(
            user_id=current_user.id,
            filename=filename,
            source_url=f"/api/v1/upload/document/{filename}",
            raw_text=extracted_text,
            structured_items=json.dumps(structured_items, ensure_ascii=False),
            summary=summary,
        )
        db.add(report)
        await db.flush()

        return {
            "status": "success",
            "filename": filename,
            "url": f"/api/v1/upload/document/{filename}",
            "size": len(content),
            "type": file.content_type,
            "extracted_text": extracted_text,
            "extraction_status": extraction_status,
            "report_id": report.id,
            "lab_summary": summary,
            "lab_items": structured_items,
        }

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"文件上传失败: {str(e)}",
        )


@router.get("/document/{filename}")
async def get_document(
    filename: str,
    current_user: User = Depends(get_current_user_required),
):
    user_dir = UPLOAD_DIR / str(current_user.id) / "documents"
    safe_filename = _safe_name(filename)
    path = user_dir / safe_filename
    if not path.exists() or not path.is_file():
        raise HTTPException(status_code=404, detail="文件不存在")
    return FileResponse(path=path, filename=safe_filename)


@router.post("/avatar")
@limiter.limit("10/minute")
async def upload_avatar(
    request: Request,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user_required),
):
    """Upload user avatar image."""
    original_name = _safe_name(file.filename or "avatar.jpg")
    ext = _extension(original_name)
    if ext not in ALLOWED_IMAGE_EXTENSIONS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="头像类型不支持，仅支持 JPG/PNG/WebP",
        )

    if file.content_type and file.content_type not in ALLOWED_IMAGE_TYPES:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="头像内容类型不支持，仅支持 JPG/PNG/WebP",
        )

    content = await file.read()
    if len(content) > MAX_IMAGE_SIZE:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"头像超过 {MAX_IMAGE_SIZE // 1024 // 1024}MB 限制",
        )

    try:
        user_dir = UPLOAD_DIR / str(current_user.id) / "avatars"
        user_dir.mkdir(parents=True, exist_ok=True)

        for existing_file in user_dir.iterdir():
            if existing_file.is_file():
                existing_file.unlink()

        filename = f"avatar{ext}"
        filepath = user_dir / filename

        with open(filepath, "wb") as f:
            f.write(content)

        prefs = _safe_json_loads(current_user.preferences)
        prefs["avatar_url"] = f"/api/v1/upload/avatar/{current_user.id}/{filename}"
        current_user.preferences = json.dumps(prefs, ensure_ascii=False)

        await db.flush()
        await db.commit()
        await db.refresh(current_user)
        
        return {
            "status": "success",
            "filename": filename,
            "url": f"/api/v1/upload/avatar/{current_user.id}/{filename}",
        }

    except Exception as e:
        await db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"头像上传失败: {str(e)}",
        )


@router.get("/avatar/{user_id}/{filename}")
async def get_avatar(
    user_id: str,
    filename: str,
):
    """Serve avatar image. Public endpoint — URL acts as capability token via user_id path.

    Note: This is intentionally public. The user_id in the path acts as a capability token.
    Do not add authentication to this endpoint.
    """
    # Prevent path traversal: only allow simple alphanumeric / UUID values
    safe_user_id = Path(user_id).name
    if not safe_user_id or safe_user_id in (".", ".."):
        raise HTTPException(status_code=400, detail="无效的用户ID")
    safe_filename = _safe_name(filename)
    user_dir = UPLOAD_DIR / safe_user_id / "avatars"
    path = user_dir / safe_filename
    if not path.exists() or not path.is_file():
        raise HTTPException(status_code=404, detail="头像不存在")
    return FileResponse(path=path, filename=safe_filename)


ALLOWED_AUDIO_EXTENSIONS = {".webm", ".ogg", ".wav", ".mp3", ".m4a", ".mp4"}
MAX_AUDIO_SIZE = 25 * 1024 * 1024  # 25MB (Whisper API limit)

ALLOWED_VIDEO_EXTENSIONS = {".mp4", ".avi", ".mov", ".mkv", ".webm"}
ALLOWED_VIDEO_TYPES = {"video/mp4", "video/avi", "video/quicktime", "video/x-matroska", "video/webm"}
MAX_VIDEO_SIZE = 100 * 1024 * 1024  # 100MB


@router.post("/audio")
@limiter.limit("20/minute")
async def upload_audio(
    request: Request,
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user_required),
):
    """Upload audio for Whisper transcription."""
    original_name = _safe_name(file.filename or "audio.webm")
    ext = _extension(original_name)
    if ext not in ALLOWED_AUDIO_EXTENSIONS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="音频格式不支持，仅支持 webm/ogg/wav/mp3/m4a",
        )

    content = await file.read()
    if len(content) > MAX_AUDIO_SIZE:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"音频文件超过 {MAX_AUDIO_SIZE // 1024 // 1024}MB 限制",
        )

    try:
        import base64 as _b64
        from openai import AsyncOpenAI

        asr_base_url = settings.WHISPER_BASE_URL or settings.OPENAI_BASE_URL
        client = AsyncOpenAI(
            api_key=settings.OPENAI_API_KEY,
            base_url=asr_base_url,
        )

        # MiMo ASR uses chat completions with input_audio, not audio.transcriptions
        mime_map = {
            ".webm": "audio/webm", ".ogg": "audio/ogg", ".wav": "audio/wav",
            ".mp3": "audio/mpeg", ".m4a": "audio/mp4", ".mp4": "audio/mp4",
        }
        mime = mime_map.get(ext, "audio/webm")
        audio_data_url = f"data:{mime};base64,{_b64.b64encode(content).decode()}"

        response = await client.chat.completions.create(
            model=settings.MIMO_ASR_MODEL,
            messages=[{
                "role": "user",
                "content": [{
                    "type": "input_audio",
                    "input_audio": {"data": audio_data_url, "format": ext.lstrip(".")}
                }]
            }],
        )

        text = response.choices[0].message.content or ""
        return {
            "status": "success",
            "text": text,
            "filename": original_name,
        }

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"语音转写失败: {str(e)}",
        )


@router.post("/video")
@limiter.limit("5/minute")
async def upload_video(
    request: Request,
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user_required),
):
    """Upload video for multimodal consultation (MiMo Omni)."""
    original_name = _safe_name(file.filename or "video.mp4")
    ext = _extension(original_name)
    if ext not in ALLOWED_VIDEO_EXTENSIONS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="视频格式不支持，仅支持 mp4/avi/mov/mkv/webm",
        )

    content = await file.read()
    if len(content) > MAX_VIDEO_SIZE:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"视频文件超过 {MAX_VIDEO_SIZE // 1024 // 1024}MB 限制",
        )

    try:
        user_dir = UPLOAD_DIR / str(current_user.id) / "videos"
        user_dir.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        filename = f"{timestamp}_{secrets.token_hex(4)}_{original_name}"
        filepath = user_dir / filename

        with open(filepath, "wb") as f:
            f.write(content)

        return {
            "status": "success",
            "filename": filename,
            "url": f"/api/v1/upload/video/{filename}",
            "size": len(content),
            "type": file.content_type,
        }

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"视频上传失败: {str(e)}",
        )


@router.get("/video/{filename}")
async def get_video(
    filename: str,
    current_user: User = Depends(get_current_user_required),
):
    """Serve uploaded video file."""
    user_dir = UPLOAD_DIR / str(current_user.id) / "videos"
    safe_filename = _safe_name(filename)
    path = user_dir / safe_filename
    if not path.exists() or not path.is_file():
        raise HTTPException(status_code=404, detail="视频文件不存在")
    return FileResponse(path=path, filename=safe_filename, media_type="video/mp4")
